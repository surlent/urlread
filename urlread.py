import urllib.request
import urllib.parse
import re
import urllib.request, urllib.parse, http.cookiejar
import openpyxl
import time
import os

def getHtml(url):
    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    opener.addheaders = [('User-Agent',
                          'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2272.101 Safari/537.36'),
                         ('Cookie', '4564564564564564565646540')]
    urllib.request.install_opener(opener)
    html_bytes = urllib.request.urlopen(url).read()
    html_string = html_bytes.decode('utf-8')
    return html_string

file=r'c:\voteanalysis.xlsx'

def sleeptime(hour,minute,sec):
    return hour*3600 + minute*60 + sec
second = sleeptime(0,10,0)
while 1==1:
    hours=time.localtime().tm_hour
    curtime = time.strftime("%m/%d %H:%M", time.localtime())
    if os.path.exists(file) and 7<=hours<=23:
        wb = openpyxl.load_workbook(file)
        ws =wb.get_sheet_by_name('voteanalysis')
        nrows=ws.max_row
        url = r'http://www.10pinping.com/vote/startin.php?id=40602'
        try:
            htmlStr = getHtml(url)
        except:
            while htmlStr.strip()=='':
                print(curtime + '   读取失败重试')
                time.sleep(30)
                htmlStr = getHtml(url)
        daTa = re.findall('\w*</div>\r\n<div class="voteNum">\d+', htmlStr, )
        for i in range(len(daTa)):
            daTa[i] = re.sub('</div>\r\n<div class="voteNum">', ",", daTa[i])
            ws.cell(row=i + nrows+1, column=1).value = daTa[i]
            ws.cell(row=i + nrows+1, column=2).value = curtime
        wb.save(file)
        wb.close()
        print(curtime + '   读取完成')
    elif os.path.exists(file):
        print('不在更新时间段')
        time.sleep(3600)
        continue
    else:
        wb = openpyxl.Workbook()  # 创建工作簿
        ws = wb.active
        ws.title = "voteanalysis"
        ws['A1'] = '原始数据'
        ws['B1'] = '抽样时间'
        url = r'http://www.10pinping.com/vote/startin.php?id=40602'
        htmlStr = getHtml(url)
        daTa = re.findall('\w*</div>\r\n<div class="voteNum">\d+', htmlStr, )
        for i in range(len(daTa)):
            daTa[i] = re.sub('</div>\r\n<div class="voteNum">', ",", daTa[i])
            ws.cell(row=i+2, column=1).value=daTa[i]
            ws.cell(row=i+2, column=2).value = curtime
        wb.save(file)
        wb.close()
        print(curtime + '   创建完成')
    time.sleep(second)